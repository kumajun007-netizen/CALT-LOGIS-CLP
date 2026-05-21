import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import math
import os
import io
from openpyxl import load_workbook
from copy import copy
from collections import defaultdict

# --- 1. 화면 스타일 및 테마 설정 ---
st.set_page_config(page_title="CALT-LOGIS CLP System", layout="wide")

MAIN_COLOR = "#001f3f"
SUB_COLOR = "#f0f2f6"
ACCENT_COLOR = "#007bff"
ALERT_COLOR = "#e74c3c"

st.markdown(f"""
    <style>
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
    * {{ font-family: 'Pretendard', sans-serif; }}
    .main {{ background-color: #f8f9fa; }}
    .header-container {{
        background-color: {MAIN_COLOR};
        padding: 25px;
        border-radius: 10px;
        color: white;
        margin-bottom: 25px;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }}
    [data-testid="stFileUploadDropzone"] {{
        border: 2px dashed {MAIN_COLOR};
        background-color: #eaf1fb;
        padding: 40px;
        border-radius: 12px;
    }}
    [data-testid="stMetric"] {{
        background-color: white;
        padding: 15px;
        border-radius: 10px;
        border-left: 5px solid {MAIN_COLOR};
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        height: 100%;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }}
    .stButton>button {{
        width: 100%;
        font-weight: 600;
        color: white;
        background-color: {MAIN_COLOR};
        border-radius: 8px;
        border: none;
        padding: 0.5rem 1rem;
    }}
    div[data-testid="stTable"] {{ font-size: 12px !important; }}
    th {{ background-color: {SUB_COLOR} !important; color: {MAIN_COLOR}; }}
    .container-box {{
        background-color: white;
        padding: 20px;
        border-radius: 10px;
        border: 1px solid #e1e4e8;
        margin-bottom: 20px;
    }}
    .guide-table {{ font-size: 11px; width: 100%; border-collapse: collapse; }}
    .guide-table th, .guide-table td {{ border: 1px solid #ddd; padding: 5px; text-align: center; }}
    .guide-table th {{ background-color: #eee; }}
    .essential {{ color: {ALERT_COLOR}; font-weight: bold; }}
    </style>
    """, unsafe_allow_html=True)

with st.container():
    st.markdown(f"""
        <div class="header-container">
            <div style="font-size: 30px; font-weight: 800; letter-spacing: -1px;">CALT-LOGIS CLP SYSTEM</div>
            <div style="font-size: 15px; opacity: 0.8; margin-top: 5px;">Busan New Port Center | Logistics Management</div>
        </div>
    """, unsafe_allow_html=True)

def clean_num(val):
    try:
        if pd.isna(val) or str(val).strip() in ['', '.', 'X', 'x', 'NaN', 'nan']: return 0.0
        return float(str(val).replace(',', '').strip())
    except: return 0.0

# --- 2. 사이드바: 엑셀 열 가이드 및 설정 ---
with st.sidebar:
    logo_path = "칼트로지스로고.png"
    if os.path.exists(logo_path):
        st.image(logo_path, use_column_width=True)

    with st.expander("📄 엑셀 업로드 표준 규격 (열 고정)", expanded=True):
        st.markdown(f"""
        <table class="guide-table">
            <tr><th>항목</th><th>엑셀 열</th><th>구분</th></tr>
            <tr class="essential"><td>No.of PKG</td><td>B 열</td><td>[필수]</td></tr>
            <tr class="essential"><td>Q'ty</td><td>F 열</td><td>[필수]</td></tr>
            <tr class="essential"><td>LENGTH (L)</td><td>J 열</td><td>[필수]</td></tr>
            <tr class="essential"><td>WIDTH (W)</td><td>L 열</td><td>[필수]</td></tr>
            <tr class="essential"><td>HEIGHT (H)</td><td>N 열</td><td>[필수]</td></tr>
            <tr><td>NET WEIGHT</td><td>I 열</td><td>선택</td></tr>
            <tr><td>REMARK (2단/3단)</td><td>O 열</td><td>선택</td></tr>
            <tr><td>LOAD (FORK_L/W/4WAY)</td><td>P 열</td><td>선택</td></tr>
            <tr><td>ITEM</td><td>D 열</td><td>참고</td></tr>
            <tr><td>DESCRIPTION</td><td>E 열</td><td>참고</td></tr>
        </table>
        <p style='font-size:11px; color:gray; margin-top:10px;'>* 데이터는 6번째 줄부터 시작하는 것을 권장합니다.</p>
        """, unsafe_allow_html=True)

    st.markdown("---")
    template_data = {
        "Invoice No": [""],
        "No.of PKG": ["PKG-001"],
        "LOCATION": [""],
        "ITEM": ["SAMPLE ITEM"],
        "Description of Goods": ["DETAIL DESC"],
        "Q'ty": [1],
        "UNIT": ["EA"],
        "Net Weight (kg)": [500],
        "Gross Weight (kg)": [550],
        "Dimension L (mm)": [1200],
        "X1": ["X"],
        "Dimension W (mm)": [1000],
        "X2": ["X"],
        "Dimension H (mm)": [2300],
        "REMARK": ["BOX,2단"],
        "LOAD": [""]
    }
    df_template = pd.DataFrame(template_data)
    tow = io.BytesIO()
    df_template.to_excel(tow, index=False, header=True, engine='openpyxl')
    st.download_button(
        label="📥 신규 화주용 양식 다운로드",
        data=tow.getvalue(),
        file_name="CALT_CLP_TEMPLATE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.header("⚙️ 배정 옵션 설정")
    with st.expander("⚖️ 컨테이너 제원", expanded=True):
        max_40_wt = st.number_input("40ft 중량 (kg)", 20000, 40000, 29500)
        max_40_len = st.number_input("40ft 길이 (mm)", 11000, 13000, 12034)
        st.markdown("---")
        max_20_wt = st.number_input("20ft 중량 (kg)", 15000, 40000, 28250)
        max_20_len = st.number_input("20ft 길이 (mm)", 5500, 6500, 5899)
        st.markdown("---")
        max_dry_h = st.number_input("DRY 높이 (mm)", 2000, 3000, 2390)
        max_hc_h = st.number_input("HC 높이 (mm)", 2000, 3500, 2695)
        max_width = st.number_input("컨테이너 폭 (mm)", 2200, 2500, 2350)

    with st.expander("🛠 적재 로직", expanded=True):
        use_balancing = st.checkbox("⚖️ 균분적재 (Balancing) ※ 끄면 컨테이너 수 최소화", value=False)
        balancing_eff = st.slider("균분적재 면적 효율 가정 (%)", 70, 100, 95, 1,
                                  help="높일수록 컨테이너 수 최소화. 낮추면 여유분 확보.")
        allow_stacking = st.checkbox("🏢 다단적재 (REMARK 2단/3단 반영)", value=True)
        fr_combine = st.checkbox("🚂 FR 혼적 허용 (같은 FR에 여러 화물 적재)", value=True,
                                 help="ON: FR 컨테이너 1대에 여러 OW/OH/OL 화물을 길이 합산해 적재. "
                                      "OFF: 화물 1개당 FR 1대 (안전·화주 분리 요구 시).")
        force_container_type = st.selectbox(
            "🏗 컨테이너 타입 (패킹 가용 한계)",
            ["자동 (높이 기준)", "40ft HC 강제", "40ft DRY 강제"],
            index=1,  # 기본값 HC 강제
            help="패킹 계산 시 사용 가능한 최대 높이 결정. "
                 "자동: 누적 H가 DRY 한계(2390) 초과 시 HC. "
                 "강제: 화주 요청 등 실무 케이스. "
                 "최종 라벨은 아래 다운그레이드 옵션이 적용됨."
        )
        auto_downgrade = st.checkbox(
            "📦 사이즈 자동 다운그레이드 (실제 사용량 기준 라벨링)",
            value=True,
            help="ON: 패킹은 40ft로 하더라도, 실제 사용 길이/높이가 작으면 20ft Dry/40ft Dry로 자동 라벨링. "
                 "화주 보고·운임 산정 정확도 향상. "
                 "OFF: 위 패킹 가용 한계 옵션 그대로 라벨링."
        )
        global_load_mode = st.selectbox(
            "🚜 전체 LOAD 모드 (개별 P열 키워드 우선)",
            ["일반 (폭 활용 최대화)",
             "FORK_W (박스W면 포크진입 → 박스W가 컨테이너 폭 방향)",
             "FORK_L (박스L면 포크진입 → 박스L이 컨테이너 폭 방향)",
             "4WAY (자유 회전)"],
            index=2,  # 기본값 FORK_L (사용자 케이스)
            help="실무 기준: 포크 진입면이 컨테이너 입구를 향함. "
                 "FORK_L=박스의 L치수가 컨테이너 폭(2350) 쪽으로, "
                 "FORK_W=박스의 W치수가 컨테이너 폭 쪽으로 적재. "
                 "박스 L > 2350인 경우 자동 회전 fallback 적용 (P열에 FORK_L_STRICT 입력 시 fallback 차단)."
        )

    if st.button("🔄 AI 재계산 실행"):
        st.session_state['manual_mode'] = False

# ===================================================================
# --- REMARK / LOAD 파싱 ---
# ===================================================================
def parse_remark(remark_str):
    """O열 REMARK 파싱: BOX, 2단, 3단"""
    if remark_str is None or pd.isna(remark_str):
        return {'layers': 1, 'is_box': True}
    s = str(remark_str).strip()
    layers = 1
    if '3단' in s: layers = 3
    elif '2단' in s: layers = 2
    is_box = 'BOX' in s.upper() or layers > 1 or s == ''
    return {'layers': layers, 'is_box': is_box}

def parse_load(load_str, global_mode):
    """
    P열 LOAD 파싱: FORK_L / FORK_W / FORK_L_STRICT / 4WAY / FR_FORCE.
    개별 키워드 우선, 없으면 전체 모드.

    - FORK_L: 박스 L → 컨테이너 폭. 단 박스 L > 컨테이너 폭이면 자동 회전 fallback.
    - FORK_L_STRICT: FORK_L 강제. fallback 차단. 박스 L > 컨테이너 폭이면 FR로 강제 이동.
    - FORK_W: 박스 W → 컨테이너 폭. fallback 동일 규칙.
    - 4WAY: 4방향 자유. 폭 활용 최대화 회전.
    - FR_FORCE: 자동 분류 무시하고 FR로 강제 이동.
    """
    if load_str is not None and not pd.isna(load_str):
        s = str(load_str).upper().strip()
        if 'FR_FORCE' in s: return 'FR_FORCE'
        if 'FORK_L_STRICT' in s: return 'FORK_L_STRICT'
        if 'FORK_L' in s: return 'FORK_L'
        if 'FORK_W' in s: return 'FORK_W'
        if '4WAY' in s: return '4WAY'
    # 전체 모드 적용 — selectbox 라벨에 키워드가 포함되어 있으므로 부분일치로 검사
    g = str(global_mode).upper()
    if 'FORK_L' in g: return 'FORK_L'
    if 'FORK_W' in g: return 'FORK_W'
    if '4WAY' in g: return '4WAY'
    return None  # 일반

def apply_rotation(l_raw, w_raw, load_mode, max_w_container):
    """
    LOAD 키워드 기준으로 박스 회전 결정.
    반환: (L_fit, W_fit, status)
      - L_fit: 컨테이너 길이방향 차지값
      - W_fit: 컨테이너 폭방향 차지값
      - status: 'OK' / 'FALLBACK' / 'IMPOSSIBLE'

    [실무 정의 — 칼트로지스 기준]
    포크 진입면이 컨테이너 입구를 향해야 함 → 진입면 치수는 컨테이너 폭 방향에 위치.

    [회전 우선순위]
    - FORK_L: 1순위 박스L → 폭(L≤2350일 때). 2순위 박스W → 폭 (fallback). 둘 다 초과면 IMPOSSIBLE.
    - FORK_L_STRICT: 1순위만. 박스L > 폭이면 IMPOSSIBLE (= FR로 강제 이동).
    - FORK_W: 1순위 박스W → 폭. 2순위 박스L → 폭 (fallback). 둘 다 초과면 IMPOSSIBLE.
    - 4WAY/일반: 폭 활용 최대화 (긴쪽이 폭에 들어가면 폭에).
    """
    if load_mode == 'FORK_L_STRICT':
        # 강제 FORK_L, fallback 차단
        if l_raw <= max_w_container:
            return w_raw, l_raw, 'OK'
        else:
            return l_raw, w_raw, 'IMPOSSIBLE'  # → FR 처리

    if load_mode == 'FORK_L':
        # 1순위: 박스 L을 컨테이너 폭에
        if l_raw <= max_w_container:
            return w_raw, l_raw, 'OK'
        # 2순위 (fallback): 박스 W를 컨테이너 폭에
        elif w_raw <= max_w_container:
            return l_raw, w_raw, 'FALLBACK'
        else:
            return l_raw, w_raw, 'IMPOSSIBLE'

    if load_mode == 'FORK_W':
        if w_raw <= max_w_container:
            return l_raw, w_raw, 'OK'
        elif l_raw <= max_w_container:
            return w_raw, l_raw, 'FALLBACK'
        else:
            return l_raw, w_raw, 'IMPOSSIBLE'

    # 4WAY / 일반: 폭 활용 최대화
    if max(l_raw, w_raw) <= max_w_container:
        return min(l_raw, w_raw), max(l_raw, w_raw), 'OK'
    elif min(l_raw, w_raw) <= max_w_container:
        return max(l_raw, w_raw), min(l_raw, w_raw), 'OK'
    else:
        return l_raw, w_raw, 'IMPOSSIBLE'

# ===================================================================
# --- 화물 분류 (STANDARD / FR) ---
# ===================================================================
def classify_cargo(piece, max_w_container, max_hc_h, max_40_len, allow_stack):
    """
    화물 1개를 분류: 'STANDARD' or 'FR'
    판정 순서 (실무 기준 그대로 적용, 임의 마진 없음):

    1. P열 'FR_FORCE' → FR (실무자 강제)
    2. 누적 H (단수×박스H) > HC 한계(2695) → OH → FR
    3. min(L, W) > 컨테이너 폭(2350) → 회전해도 OW → FR
    4. max(L, W) > 40ft L 한계(12034) → OL → FR
    5. 그 외 → STANDARD
    """
    l, w, h = piece['L_raw'], piece['W_raw'], piece['H_raw']
    layers = piece['layers'] if allow_stack else 1
    total_h = h * layers
    load_mode = piece.get('load_mode')

    # 1. FR 강제
    if load_mode == 'FR_FORCE':
        return 'FR', 'FR_FORCE'

    # 2. OH (다단 후 누적)
    if total_h > max_hc_h:
        return 'FR', 'OH'

    # 3. OW (회전 무관)
    if min(l, w) > max_w_container:
        return 'FR', 'OW'

    # 4. OL (회전 무관)
    if max(l, w) > max_40_len:
        return 'FR', 'OL'

    return 'STANDARD', None

# ===================================================================
# --- 다단 묶음 → 풋프린트 단위 생성 (STANDARD / FR 분리) ---
# ===================================================================
def build_footprints(df, max_w_container, max_hc_h, max_40_len, allow_stack):
    """
    화물을 STANDARD/FR로 분류 후 각각 풋프린트 생성.

    [STANDARD 풀]
    - 같은 (PKG, 치수, 단수, LOAD) 박스를 N단씩 묶어 풋프린트 단위로 만듦.
    - apply_rotation 적용. status가 'IMPOSSIBLE'이면 FR로 재분류.

    [FR 풀]
    - 회전 없음 (박스 L 그대로 컨테이너 L 방향, 사용자 정의)
    - 다단 적용 없음 (FR 화물은 단일 적재 원칙)
    - 박스 1개당 풋프린트 1개

    반환: (standard_fps, fr_fps, classify_summary)
    """
    standard_pieces = []
    fr_pieces = []
    classify_summary = defaultdict(int)

    # 1차 분류
    for _, row in df.iterrows():
        piece = row.to_dict()
        cargo_type, reason = classify_cargo(piece, max_w_container, max_hc_h, max_40_len, allow_stack)
        piece['fr_reason'] = reason
        if cargo_type == 'FR':
            fr_pieces.append(piece)
            classify_summary[f"FR ({reason})"] += 1
        else:
            standard_pieces.append(piece)
            classify_summary['STANDARD'] += 1

    # STANDARD 풀 풋프린트 생성
    standard_fps = []
    grouped = defaultdict(list)
    for r in standard_pieces:
        layers_use = r['layers'] if allow_stack else 1
        key = (r['PKG NO'], int(r['L_raw']), int(r['W_raw']), int(r['H_raw']),
               layers_use, r['load_mode'])
        grouped[key].append(r)

    for key, items in grouped.items():
        pkg, l_raw, w_raw, h_raw, layers, load_mode = key
        i = 0
        while i < len(items):
            stack = items[i:i+layers]
            i += layers
            n = len(stack)
            l_fit, w_fit, rot_status = apply_rotation(l_raw, w_raw, load_mode, max_w_container)

            # 회전 불가 → FR로 재분류
            if rot_status == 'IMPOSSIBLE':
                for s in stack:
                    s['fr_reason'] = 'ROT_IMPOSSIBLE'
                    fr_pieces.append(s)
                    classify_summary[f"FR (ROT_IMPOSSIBLE)"] += 1
                    classify_summary['STANDARD'] -= 1
                continue

            fp = {
                'PKG NO': pkg,
                'ITEM': stack[0].get('ITEM', '-'),
                'GROUP': stack[0].get('GROUP', '-'),
                'L': l_fit,
                'W': w_fit,
                'H': h_raw * n,
                'H_per_layer': h_raw,
                'stack_count': n,
                'WEIGHT': sum(s['WEIGHT'] for s in stack),
                'load_mode': load_mode if load_mode else '-',
                'rotation_status': rot_status,
                'STACK_OK': n > 1,
                'row_idx': stack[0]['row_idx'],
                'sub_items': stack,
                'cargo_type': 'STANDARD',
            }
            standard_fps.append(fp)

    # FR 풀 풋프린트 생성 (회전 없음)
    fr_fps = []
    for piece in fr_pieces:
        l, w, h = int(piece['L_raw']), int(piece['W_raw']), int(piece['H_raw'])
        # FR: 박스 L → 컨테이너 L 그대로 (사용자 정의: 옆에서 적재하므로 L축 정렬)
        fp = {
            'PKG NO': piece['PKG NO'],
            'ITEM': piece.get('ITEM', '-'),
            'GROUP': piece.get('GROUP', '-'),
            'L': l,  # 컨테이너 길이 방향 = 박스 L
            'W': w,  # 컨테이너 폭 방향 = 박스 W (OW 가능)
            'H': h,
            'H_per_layer': h,
            'stack_count': 1,
            'WEIGHT': int(piece['WEIGHT']),
            'load_mode': piece.get('load_mode') if piece.get('load_mode') else '-',
            'rotation_status': 'FR_NO_ROTATE',
            'STACK_OK': False,
            'row_idx': piece['row_idx'],
            'sub_items': [piece],
            'cargo_type': 'FR',
            'fr_reason': piece.get('fr_reason', '-'),
        }
        fr_fps.append(fp)

    return standard_fps, fr_fps, dict(classify_summary)

# ===================================================================
# --- 컨테이너 타입 사전 선정 ---
# ===================================================================
def decide_container_height(footprints, max_dry_h, max_hc_h, force_type='자동 (높이 기준)'):
    """전체 풋프린트의 최대 누적 높이로 HC/Dry 선정. 강제 옵션 우선."""
    if 'HC' in force_type:
        return max_hc_h, 'HC'
    if 'DRY' in force_type:
        return max_dry_h, 'DRY'
    # 자동
    if not footprints:
        return max_dry_h, 'DRY'
    max_h = max(fp['H'] for fp in footprints)
    if max_h > max_dry_h:
        return max_hc_h, 'HC'
    else:
        return max_dry_h, 'DRY'

# ===================================================================
# --- 풋프린트 패킹 로직 (Maximal Rectangles + BSSF) ---
# ===================================================================
def _prune_contained_rects(rects):
    """포함 관계인 사각형 제거 (다른 rect 안에 완전히 들어가면 제거).
    free_rects 누적 비효율 방지 — Maximal Rectangles 핵심."""
    cleaned = []
    n = len(rects)
    for i in range(n):
        ax, ay, aw, ah = rects[i]['x'], rects[i]['y'], rects[i]['w'], rects[i]['h']
        ax2, ay2 = ax + aw, ay + ah
        contained = False
        for j in range(n):
            if i == j: continue
            bx, by, bw, bh = rects[j]['x'], rects[j]['y'], rects[j]['w'], rects[j]['h']
            bx2, by2 = bx + bw, by + bh
            if (bx <= ax and by <= ay and bx2 >= ax2 and by2 >= ay2
                and (bw > aw or bh > ah or (bw == aw and bh == ah and i > j))):
                contained = True
                break
        if not contained:
            cleaned.append(rects[i])
    return cleaned

def pack_footprint_into_bin(fp, b, max_wt, max_len, max_h_container, max_w_container):
    """
    Maximal Rectangles 기반 2D 패킹 (BSSF: Best Short Side Fit).

    [원리]
    - 컨테이너 바닥을 빈 사각형 리스트(free_rects)로 관리
    - 새 박스가 들어갈 때 가장 딱 맞는 빈 사각형 선택 (short side 잔여공간 최소)
    - 배치된 박스와 겹치는 모든 free_rect를 4방향(L/R/U/D)으로 분할
    - 포함 관계인 사각형은 즉시 제거하여 free_rects 누적 방지
    - 작은 박스가 큰 박스의 빈 옆/뒷자리에 자연스럽게 끼워 들어감

    [좌표계]
    - x축: 컨테이너 L방향 (0 ~ max_len)
    - y축: 컨테이너 W방향 (0 ~ max_w_container)
    """
    # 사전 체크
    if b['total_W'] + fp['WEIGHT'] > max_wt: return False
    if fp['H'] > max_h_container: return False
    if fp['L'] > max_len: return False
    if fp['W'] > max_w_container: return False

    # 빈 사각형 / 배치 리스트 초기화
    if 'free_rects' not in b:
        b['free_rects'] = [{'x': 0, 'y': 0, 'w': max_len, 'h': max_w_container}]
        b['placed'] = []

    # BSSF: 가장 적합한 빈 사각형 찾기
    best_idx = -1
    best_short = float('inf')
    best_long = float('inf')
    for i, r in enumerate(b['free_rects']):
        if fp['L'] <= r['w'] and fp['W'] <= r['h']:
            leftover_short = min(r['w'] - fp['L'], r['h'] - fp['W'])
            leftover_long = max(r['w'] - fp['L'], r['h'] - fp['W'])
            if (leftover_short < best_short
                or (leftover_short == best_short and leftover_long < best_long)):
                best_short = leftover_short
                best_long = leftover_long
                best_idx = i

    if best_idx == -1:
        return False

    # 배치 좌표 결정
    target = b['free_rects'][best_idx]
    placed_fp = dict(fp)
    placed_fp['x'] = target['x']
    placed_fp['y'] = target['y']
    b['placed'].append(placed_fp)

    # 통계 업데이트
    new_x_end = placed_fp['x'] + fp['L']
    new_y_end = placed_fp['y'] + fp['W']
    b['used_L'] = max(b['used_L'], new_x_end)
    b['max_W'] = max(b['max_W'], new_y_end)
    b['max_H'] = max(b['max_H'], fp['H'])
    b['total_W'] += fp['WEIGHT']
    b['groups'].add(fp['GROUP'])

    # Maximal Rectangles: 배치된 박스와 겹치는 모든 free_rect를 4방향 분할
    used_x1, used_y1 = placed_fp['x'], placed_fp['y']
    used_x2, used_y2 = new_x_end, new_y_end

    new_rects = []
    for fr in b['free_rects']:
        fx, fy, fw, fh = fr['x'], fr['y'], fr['w'], fr['h']
        fx2, fy2 = fx + fw, fy + fh
        # 사용 영역과 겹치지 않으면 그대로 유지
        if used_x1 >= fx2 or used_x2 <= fx or used_y1 >= fy2 or used_y2 <= fy:
            new_rects.append(fr)
            continue
        # 겹침 → 4방향(왼/오/아래/위)으로 분할
        # 왼쪽 영역
        if used_x1 > fx:
            new_rects.append({'x': fx, 'y': fy, 'w': used_x1 - fx, 'h': fh})
        # 오른쪽 영역
        if used_x2 < fx2:
            new_rects.append({'x': used_x2, 'y': fy, 'w': fx2 - used_x2, 'h': fh})
        # 아래 영역
        if used_y1 > fy:
            new_rects.append({'x': fx, 'y': fy, 'w': fw, 'h': used_y1 - fy})
        # 위 영역
        if used_y2 < fy2:
            new_rects.append({'x': fx, 'y': used_y2, 'w': fw, 'h': fy2 - used_y2})

    # 포함 관계 제거 (Maximal Rectangles 핵심)
    b['free_rects'] = _prune_contained_rects(new_rects)

    # 호환성 유지를 위한 rows 구조 (시각화는 placed 사용, rows는 다운로드 매핑 등에 사용)
    b['rows'].append({'items': [placed_fp], 'used_W': fp['W'], 'max_L': fp['L']})

    return True

# ===================================================================
# --- 라벨 적용 (컨테이너별 실제 사용량 기준 자동 다운사이징) ---
# ===================================================================
def apply_labels(bins, max_20_len, max_20_wt, fr_max_len, max_dry_h, max_hc_h,
                 max_w_container, container_h_used, auto_downgrade=True):
    """
    각 컨테이너의 실제 사용량(max_H, used_L, total_W)을 기준으로 라벨링.

    [auto_downgrade=True - 기본] 실제 사용량 기준 자동 다운사이징
      - STANDARD: max_H 기준으로 HC/Dry, used_L+total_W 기준으로 20ft/40ft 선택
      - FR: used_L 기준으로 20ft/40ft Flat Rack 선택

    [auto_downgrade=False] 패킹 가용 한계 그대로 라벨링
      - container_h_used가 HC면 모두 40ft HC, DRY면 길이 기준 20ft/40ft Dry
    """
    for b in bins:
        is_fr_bin = b.get('is_fr', False)

        if is_fr_bin:
            # FR 라벨링 — auto_downgrade와 무관하게 used_L 기준 다운사이징
            is_20ft_size = b['used_L'] <= max_20_len and b['total_W'] <= max_20_wt
            tags = []
            if b['max_H'] > max_hc_h: tags.append(f"OH+{b['max_H']-max_hc_h}")
            if b['max_W'] > max_w_container: tags.append(f"OW+{b['max_W']-max_w_container}")
            if b['used_L'] > fr_max_len: tags.append(f"OL+{b['used_L']-fr_max_len}")
            if not tags:
                tags.append("FR")
            base = "20ft Flat Rack" if is_20ft_size else "40ft Flat Rack"
            b['c_label'] = f"{base} [{' + '.join(tags)}] #{b['id']}"
        else:
            is_20ft_size = (b['used_L'] <= max_20_len
                            and b['total_W'] <= max_20_wt)

            if auto_downgrade:
                # 실제 사용 max_H 기준
                needs_hc = b['max_H'] > max_dry_h
            else:
                # 패킹 가용 한계 그대로
                needs_hc = (container_h_used == 'HC') or (b['max_H'] > max_dry_h)

            if needs_hc:
                base = "40ft HC"  # 20ft HC는 실무에서 드물어 제외
            else:
                base = "20ft Dry" if is_20ft_size else "40ft Dry"
            b['c_label'] = f"{base} #{b['id']}"
    return bins

# ===================================================================
# --- FR 전용 패킹 (길이 합산만 체크, 회전 없음) ---
# ===================================================================
def pack_fr_into_bin(fp, b, max_len, max_wt):
    """
    FR 컨테이너에 화물 적재.
    - 길이 합산만 체크 (폭/높이는 OW/OH 태그로 통과)
    - 화물 회전 없음 (사용자 정의: FR은 옆에서 적재, 박스 L 그대로)
    - 중량 합산은 체크 (40ft FR 중량 한계 적용)
    """
    if b['used_L'] + fp['L'] > max_len: return False
    if b['total_W'] + fp['WEIGHT'] > max_wt: return False
    b['rows'].append({'items': [fp], 'used_W': fp['W'], 'max_L': fp['L']})
    b['used_L'] += fp['L']
    b['total_W'] += fp['WEIGHT']
    b['max_W'] = max(b['max_W'], fp['W'])
    b['max_H'] = max(b['max_H'], fp['H'])
    b['groups'].add(fp['GROUP'])
    return True

# ===================================================================
# --- 메인 패킹 계산 (STANDARD + FR 분리 처리) ---
# ===================================================================
def calculate_expert_packing(df, max_40_wt, max_40_len, max_20_wt, max_20_len,
                             max_dry_h, max_hc_h, max_w_container,
                             use_balancing, allow_stack, force_container_type,
                             global_load_mode, balancing_eff=95, fr_combine=True,
                             auto_downgrade=True):
    # 1. 풋프린트 생성 (STANDARD/FR 분리)
    standard_fps, fr_fps, classify_summary = build_footprints(
        df, max_w_container, max_hc_h, max_40_len, allow_stack
    )

    # 2. STANDARD 컨테이너 타입 결정 (강제 옵션 우선)
    max_h_container, h_type = decide_container_height(
        standard_fps, max_dry_h, max_hc_h, force_container_type
    )

    # 3. STANDARD 풋프린트 정렬: W 큰순 → H 큰순 → L 큰순 → GROUP
    # 3. STANDARD 풋프린트 정렬 — Maximal Rectangles 효율을 위해 가장 긴 변 우선
    #    1순위: max(L, W) 큰 순 (막대형 화물이 자리 잡지 못해 빈 컨테이너로 빠지는 현상 방지)
    #    2순위: 면적(L×W) 큰 순
    #    3순위: H 큰 순
    #    4순위: GROUP (같은 PKG 인접 배치)
    standard_fps.sort(key=lambda x: (-max(x['L'], x['W']), -(x['L']*x['W']), -x['H'], str(x['GROUP'])))

    standard_bins = []
    c_no = 1

    # 4. 균분적재용 사전 컨테이너 생성 (STANDARD 풀만)
    if use_balancing and standard_fps:
        total_w_kg = sum(fp['WEIGHT'] for fp in standard_fps)
        floor_area_needed = sum(fp['L'] * fp['W'] for fp in standard_fps)
        floor_area_per_bin = max_40_len * max_w_container * (balancing_eff / 100.0)
        est_by_area = math.ceil(floor_area_needed / floor_area_per_bin)
        est_by_wt = math.ceil(total_w_kg / max_40_wt) if total_w_kg > 0 else 1
        est_bins = max(1, est_by_area, est_by_wt)
        for _ in range(est_bins):
            standard_bins.append({
                'id': c_no, 'rows': [], 'used_L': 0, 'total_W': 0,
                'max_W': 0, 'max_H': 0, 'groups': set(), 'is_fr': False
            })
            c_no += 1

    # 5. STANDARD 풋프린트 배치
    for fp in standard_fps:
        if use_balancing and standard_bins:
            standard_bins.sort(key=lambda b: (0 if fp['GROUP'] in b['groups'] else 1, b['used_L']))

        placed = False
        for b in standard_bins:
            if pack_footprint_into_bin(fp, b, max_40_wt, max_40_len, max_h_container, max_w_container):
                placed = True
                break

        if not placed:
            new_bin = {
                'id': c_no, 'rows': [], 'used_L': 0, 'total_W': 0,
                'max_W': 0, 'max_H': 0, 'groups': set(), 'is_fr': False
            }
            if pack_footprint_into_bin(fp, new_bin, max_40_wt, max_40_len, max_h_container, max_w_container):
                standard_bins.append(new_bin)
                c_no += 1
            else:
                # 어떻게도 안 들어가는 풋프린트 → FR로 이동
                fr_fps.append({**fp, 'cargo_type': 'FR', 'fr_reason': 'OVERFLOW',
                               'rotation_status': 'FR_NO_ROTATE'})

    # 6. FR 풋프린트 정렬: L 긴 순 → W 큰 순 (긴 화물 먼저 배치)
    fr_fps.sort(key=lambda x: (-x['L'], -x['W']))

    fr_bins = []
    # 7. FR 패킹
    for fp in fr_fps:
        placed = False
        if fr_combine:
            # 같은 FR에 길이 합산해서 혼적
            for b in fr_bins:
                if pack_fr_into_bin(fp, b, max_40_len, max_40_wt):
                    placed = True
                    break
        if not placed:
            new_bin = {
                'id': c_no, 'rows': [], 'used_L': 0, 'total_W': 0,
                'max_W': 0, 'max_H': 0, 'groups': set(), 'is_fr': True
            }
            pack_fr_into_bin(fp, new_bin, max_40_len, max_40_wt)
            fr_bins.append(new_bin)
            c_no += 1

    # 8. 통합 + 빈 컨테이너 정리 + ID 재정렬
    all_bins = standard_bins + fr_bins
    all_bins = [b for b in all_bins if b['used_L'] > 0]
    # STANDARD 먼저, FR 나중에 정렬
    all_bins.sort(key=lambda x: (x.get('is_fr', False), x['id']))
    for idx, b in enumerate(all_bins):
        b['id'] = idx + 1

    return apply_labels(all_bins, max_20_len, max_20_wt, max_40_len - 430,
                        max_dry_h, max_hc_h, max_w_container, h_type,
                        auto_downgrade), classify_summary

# ===================================================================
# --- 3. 메인 화면 로직 ---
# ===================================================================
def reset_data():
    if 'bins' in st.session_state: del st.session_state['bins']
    if 'manual_mode' in st.session_state: del st.session_state['manual_mode']

st.markdown("### 📤 패킹리스트 업로드 (드래그 앤 드롭)")
file = st.file_uploader("이곳에 파일을 끌어다 놓으세요.", type=['csv', 'xlsx'], on_change=reset_data)

if file is not None:
    try:
        selected_sheet = None

        # 엑셀 고정 열 매핑 (0-base)
        COL_PKG = 1       # B열: No.of PKG
        COL_ITEM = 3      # D열: ITEM
        COL_DESC = 4      # E열: DESCRIPTION
        COL_QTY = 5       # F열: Q'ty  ★ 신규 반영
        COL_WEIGHT = 8    # I열: WEIGHT
        COL_L = 9         # J열: LENGTH
        COL_W = 11        # L열: WIDTH
        COL_H = 13        # N열: HEIGHT
        COL_REMARK = 14   # O열: REMARK ★ 신규 반영
        COL_LOAD = 15     # P열: LOAD   ★ 신규 반영

        def count_valid_rows(sheet_df):
            valid_count = 0
            for _, r in sheet_df.iterrows():
                try:
                    pkg = str(r.iloc[COL_PKG]).replace('00:00:00', '').replace('.0', '').strip() if pd.notna(r.iloc[COL_PKG]) else ""
                    l = clean_num(r.iloc[COL_L])
                    w = clean_num(r.iloc[COL_W])
                    h = clean_num(r.iloc[COL_H])
                    if pkg.lower() not in ['', 'nan', 'none', '.'] and l > 0 and w > 0 and h > 0:
                        valid_count += 1
                except:
                    continue
            return valid_count

        if file.name.lower().endswith('.xlsx'):
            file.seek(0)
            excel_sheets = pd.read_excel(file, sheet_name=None, header=None)

            raw_full = None
            best_count = 0

            for sheet_name, sheet_df in excel_sheets.items():
                valid_count = count_valid_rows(sheet_df)
                if valid_count > best_count:
                    best_count = valid_count
                    raw_full = sheet_df
                    selected_sheet = sheet_name

            if raw_full is None or best_count == 0:
                st.error("엑셀에서 B/J/L/N 열 기준의 필수 데이터를 찾지 못했습니다.")
                with st.expander("🔎 확인 필요 사항", expanded=True):
                    st.write("- B열: No.of PKG / F열: Q'ty / J열: LENGTH / L열: WIDTH / N열: HEIGHT")
                    st.write("- O열(REMARK)에 '2단','3단' 키워드 / P열(LOAD)에 FORK_L/W/4WAY 가능")
                st.stop()

            st.info(f"읽은 시트: {selected_sheet} / 인식된 화물 행 수: {best_count} PKG (Q'ty 복제 전)")

        else:
            file.seek(0)
            raw_full = pd.read_csv(file, header=None)
            selected_sheet = None

        p_data = []
        total_qty = 0

        # ★ 핵심 변경: Q'ty 만큼 박스 복제
        for orig_idx in range(len(raw_full)):
            row = raw_full.iloc[orig_idx]

            try:
                pkg_v = str(row.iloc[COL_PKG]).replace('00:00:00', '').replace('.0', '').strip() if pd.notna(row.iloc[COL_PKG]) else ""
                l_v = clean_num(row.iloc[COL_L])
                w_v = clean_num(row.iloc[COL_W])
                h_v = clean_num(row.iloc[COL_H])
                weight_v = clean_num(row.iloc[COL_WEIGHT])
                qty_v = int(clean_num(row.iloc[COL_QTY])) if row.shape[0] > COL_QTY else 1
                if qty_v <= 0: qty_v = 1

                # REMARK / LOAD 파싱
                remark_raw = row.iloc[COL_REMARK] if row.shape[0] > COL_REMARK else None
                load_raw = row.iloc[COL_LOAD] if row.shape[0] > COL_LOAD else None
                r_info = parse_remark(remark_raw)
                load_mode = parse_load(load_raw, global_load_mode)

                if pkg_v.lower() in ['', 'nan', 'none', '.'] or l_v == 0 or w_v == 0 or h_v == 0:
                    continue

                # Q'ty 만큼 복제
                for q_i in range(qty_v):
                    p_data.append({
                        'PKG NO': pkg_v,
                        'GROUP': str(row.iloc[COL_DESC]) if pd.notna(row.iloc[COL_DESC]) else "-",
                        'ITEM': str(row.iloc[COL_ITEM]) if pd.notna(row.iloc[COL_ITEM]) else "-",
                        'DESC': str(row.iloc[COL_DESC]) if pd.notna(row.iloc[COL_DESC]) else "-",
                        'L_raw': int(l_v + 0.5),
                        'W_raw': int(w_v + 0.5),
                        'H_raw': int(h_v + 0.5),
                        'WEIGHT': int(weight_v + 0.5),
                        'layers': r_info['layers'],
                        'load_mode': load_mode,
                        'row_idx': int(orig_idx),
                    })
                total_qty += qty_v
            except:
                continue

        df = pd.DataFrame(p_data)
        if df.empty:
            st.warning("⚠️ 지정된 열(B, F, J, L, N)에서 필수 데이터를 찾을 수 없습니다. 양식을 확인하세요.")
        else:
            st.success(f"✅ 총 {len(df):,} 박스 복제 완료 (Q'ty 반영)")

            if 'bins' not in st.session_state or not st.session_state.get('manual_mode', False):
                bins_result, classify_summary = calculate_expert_packing(
                    df, max_40_wt, max_40_len, max_20_wt, max_20_len,
                    max_dry_h, max_hc_h, max_width,
                    use_balancing, allow_stacking, force_container_type,
                    global_load_mode, balancing_eff, fr_combine, auto_downgrade
                )
                st.session_state.bins = bins_result
                st.session_state.classify_summary = classify_summary
                st.session_state.manual_mode = True

            bins = st.session_state.bins
            classify_summary = st.session_state.get('classify_summary', {})

            # 분류 결과 요약 표시
            if classify_summary:
                with st.expander("🔍 화물 분류 결과", expanded=False):
                    s_cnt = classify_summary.get('STANDARD', 0)
                    fr_items = {k: v for k, v in classify_summary.items() if k.startswith('FR')}
                    st.markdown(f"- **일반(STANDARD)**: {s_cnt}박스")
                    if fr_items:
                        st.markdown("- **FR(특수)**:")
                        for k, v in fr_items.items():
                            reason = k.replace('FR (', '').replace(')', '')
                            label = {
                                'OH': '높이 초과 (>2,695mm)',
                                'OW': '폭 초과 (회전해도 >2,350mm)',
                                'OL': '길이 초과 (>12,034mm)',
                                'FR_FORCE': 'P열 FR_FORCE 강제',
                                'ROT_IMPOSSIBLE': '회전 fallback 차단 (FORK_L_STRICT)',
                                'OVERFLOW': '일반 컨테이너 적재 실패 (잔여 공간 없음)'
                            }.get(reason, reason)
                            st.markdown(f"   - {label}: {v}박스")

            target_options = [f"{b_opt['id']}번" for b_opt in bins] + ["✨ 새 컨테이너"]

            st.subheader("📊 실시간 적재 요약")
            m1, m2, m3, m4 = st.columns(4)
            def _all_items(b):
                """placed 우선, 없으면 rows에서 가져옴"""
                if b.get('placed'):
                    return b['placed']
                return [fp for r in b['rows'] for fp in r['items']]
            packed_fp = sum(len(_all_items(b)) for b in bins)
            packed_boxes = sum(
                sum(fp['stack_count'] for fp in _all_items(b)) for b in bins
            )
            m1.metric("전체 화물(박스)", f"{len(df):,} PKG")
            m2.metric("배정 박스", f"{packed_boxes:,} PKG")
            m3.metric("컨테이너 수", f"{len(bins)} UNIT")
            avg_wt = sum(b['total_W'] for b in bins) / len(bins) if bins else 0
            m4.metric("평균 중량", f"{avg_wt:,.0f} kg")

            for b in bins:
                if b['used_L'] == 0: continue
                st.markdown(f'<div class="container-box">', unsafe_allow_html=True)
                st.markdown(f"### 📦 {b['c_label']}")

                # 풋프린트 상세 — placed 우선 (Maximal Rectangles 결과), 없으면 rows fallback
                t_data = []
                placed_items = b.get('placed', [])
                if placed_items:
                    for fp in placed_items:
                        t_data.append({
                            'PKG NO': fp['PKG NO'],
                            'ITEM': fp['ITEM'],
                            'L(회전후)': fp['L'],
                            'W(회전후)': fp['W'],
                            'H(누적)': fp['H'],
                            '단수': fp['stack_count'],
                            'LOAD': fp['load_mode'],
                            'WEIGHT': fp['WEIGHT'],
                            '이동': f"{b['id']}번"
                        })
                else:
                    for r in b['rows']:
                        for fp in r['items']:
                            t_data.append({
                                'PKG NO': fp['PKG NO'],
                                'ITEM': fp['ITEM'],
                                'L(회전후)': fp['L'],
                                'W(회전후)': fp['W'],
                                'H(누적)': fp['H'],
                                '단수': fp['stack_count'],
                                'LOAD': fp['load_mode'],
                                'WEIGHT': fp['WEIGHT'],
                                '이동': f"{b['id']}번"
                            })

                df_edit = pd.DataFrame(t_data)
                edited_df = st.data_editor(
                    df_edit, hide_index=True, use_container_width=True, key=f"ed_{b['id']}",
                    column_config={"이동": st.column_config.SelectboxColumn("🚚 이동", options=target_options)},
                    disabled=['PKG NO', 'ITEM', 'L(회전후)', 'W(회전후)', 'H(누적)', '단수', 'LOAD', 'WEIGHT']
                )

                with st.expander("👁️ 적재 단면도 및 제원 확인", expanded=True):
                    cur_max_l = max_20_len if "20ft" in b['c_label'] else max_40_len
                    cur_max_w = max_20_wt if "20ft" in b['c_label'] else max_40_wt
                    cur_max_h = max_hc_h if "HC" in b['c_label'] else max_dry_h
                    # placed 우선, 없으면 rows fallback
                    used_width = b.get('max_W', 0) or max([r['used_W'] for r in b['rows']] + [0])
                    used_height = b['max_H']

                    c1, c2, c3, c4 = st.columns(4)
                    c1.markdown(f"**📏 길이:** {b['used_L']:,}/{cur_max_l:,}mm")
                    c1.progress(min(1.0, b['used_L']/cur_max_l))
                    c2.markdown(f"**⚖️ 중량:** {b['total_W']:,}/{cur_max_w:,}kg")
                    c2.progress(min(1.0, b['total_W']/cur_max_w))

                    if used_width > max_width:
                        c3.markdown(f"**↔️ 폭:** {used_width:,}/{max_width:,}mm <span style='color:{ALERT_COLOR};font-weight:bold;'>[OW +{used_width-max_width:,}]</span>", unsafe_allow_html=True)
                    else:
                        c3.markdown(f"**↔️ 폭:** {used_width:,}/{max_width:,}mm")
                        c3.progress(min(1.0, used_width/max_width))

                    if used_height > cur_max_h:
                        c4.markdown(f"**↕️ 높이:** {used_height:,}/{cur_max_h:,}mm <span style='color:{ALERT_COLOR};font-weight:bold;'>[OH +{used_height-cur_max_h:,}]</span>", unsafe_allow_html=True)
                    else:
                        c4.markdown(f"**↕️ 높이:** {used_height:,}/{cur_max_h:,}mm")
                        c4.progress(min(1.0, used_height/cur_max_h))

                    # 평면도 — Maximal Rectangles 결과의 (x, y) 좌표 직접 사용
                    fig = go.Figure()
                    fig.add_shape(type="rect", x0=b['used_L'], y0=0, x1=cur_max_l, y1=max_width,
                                  fillcolor="#e1e4e8", opacity=0.4, line_width=0)
                    fig.add_shape(type="rect", x0=0, y0=0, x1=cur_max_l, y1=max_width,
                                  line=dict(color=MAIN_COLOR, width=2))

                    # placed 우선 사용 (Maximal Rectangles 결과). 없으면 rows fallback (FR 등).
                    placed_items = b.get('placed', [])
                    if placed_items:
                        for fp in placed_items:
                            x0, y0 = fp.get('x', 0), fp.get('y', 0)
                            x1, y1 = x0 + fp['L'], y0 + fp['W']
                            fig.add_shape(type="rect",
                                          x0=x0, y0=y0, x1=x1, y1=y1,
                                          fillcolor=ACCENT_COLOR, opacity=0.8,
                                          line=dict(color="white", width=1))
                            label_text = (f"{fp['PKG NO']}<br>{fp['stack_count']}단"
                                          if fp['stack_count'] > 1 else str(fp['PKG NO']))
                            fig.add_annotation(x=(x0+x1)/2, y=(y0+y1)/2,
                                               text=label_text, showarrow=False,
                                               font=dict(color="white", size=9))
                    else:
                        # FR 등 placed 없는 컨테이너 fallback
                        cx = 0
                        for r in b['rows']:
                            cy = (max_width - r['used_W']) / 2
                            for fp in r['items']:
                                fig.add_shape(type="rect",
                                              x0=cx, y0=cy, x1=cx+fp['L'], y1=cy+fp['W'],
                                              fillcolor=ACCENT_COLOR, opacity=0.8,
                                              line=dict(color="white", width=1))
                                label_text = (f"{fp['PKG NO']}<br>{fp['stack_count']}단"
                                              if fp['stack_count'] > 1 else str(fp['PKG NO']))
                                fig.add_annotation(x=cx+fp['L']/2, y=cy+fp['W']/2,
                                                   text=label_text, showarrow=False,
                                                   font=dict(color="white", size=9))
                                cy += fp['W']
                            cx += r['max_L']

                    fig.add_shape(type="line", x0=b['used_L'], y0=-200, x1=b['used_L'], y1=max_width+450,
                                  line=dict(color=ALERT_COLOR, width=2, dash="dash"))
                    if b['used_L'] > 100:
                        fig.add_annotation(x=b['used_L']/2, y=max_width+300,
                                           text=f"적재: {b['used_L']:,}mm",
                                           showarrow=False, font=dict(color=MAIN_COLOR, size=13))
                    if cur_max_l - b['used_L'] > 100:
                        fig.add_annotation(x=b['used_L'] + (cur_max_l - b['used_L'])/2, y=max_width+300,
                                           text=f"잔여: {cur_max_l - b['used_L']:,}mm",
                                           showarrow=False, font=dict(color=ALERT_COLOR, size=13))
                    fig.update_layout(xaxis=dict(visible=False, range=[-200, max_40_len+400]),
                                      yaxis=dict(visible=False, range=[-300, max_width+750]),
                                      height=280, margin=dict(l=10, r=10, t=30, b=10),
                                      paper_bgcolor="rgba(0,0,0,0)")
                    st.plotly_chart(fig, use_container_width=True, key=f"plot_{b['id']}")
                st.markdown('</div>', unsafe_allow_html=True)

            # --- 최종 결과 다운로드 ---
            st.markdown("---")

            mapping = {}
            for bx in bins:
                # placed 우선 (STANDARD), rows fallback (FR)
                items_iter = bx.get('placed') or [
                    fp for r in bx['rows'] for fp in r['items']
                ]
                for fp in items_iter:
                    for sub in fp.get('sub_items', []):
                        mapping[sub['row_idx']] = bx['c_label']

            if file.name.endswith('.xlsx'):
                file.seek(0)
                wb = load_workbook(file)
                ws = wb[selected_sheet]

                target_col = ws.max_column + 1
                target_letter = ws.cell(row=1, column=target_col).column_letter
                ws.cell(row=4, column=target_col).value = "배정 컨테이너"

                if target_col > 1:
                    src_cell = ws.cell(row=4, column=target_col - 1)
                    dst_cell = ws.cell(row=4, column=target_col)
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format

                for r_idx, label in mapping.items():
                    excel_row = int(r_idx) + 1
                    ws.cell(row=excel_row, column=target_col).value = label
                    if target_col > 1:
                        src_cell = ws.cell(row=excel_row, column=target_col - 1)
                        dst_cell = ws.cell(row=excel_row, column=target_col)
                        dst_cell.font = copy(src_cell.font)
                        dst_cell.fill = copy(src_cell.fill)
                        dst_cell.border = copy(src_cell.border)
                        dst_cell.alignment = copy(src_cell.alignment)
                        dst_cell.number_format = src_cell.number_format

                ws.column_dimensions[target_letter].width = 25

                output = io.BytesIO()
                wb.save(output)

                st.download_button(
                    label="📥 최종 결과 다운로드 (원본 양식 유지)",
                    data=output.getvalue(),
                    file_name="CLP_RESULT_FINAL.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            else:
                export_df = raw_full.copy()
                target_col = export_df.shape[1]
                export_df[target_col] = ""
                export_df.iloc[3, target_col] = "배정 컨테이너"

                for r_idx, label in mapping.items():
                    export_df.iloc[int(r_idx), target_col] = label

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, header=False)

                st.download_button(
                    label="📥 최종 결과 다운로드",
                    data=output.getvalue(),
                    file_name="CLP_RESULT_FINAL.xlsx",
                    use_container_width=True
                )

    except Exception as e:
        st.error(f"데이터 처리 중 오류 발생: {e}")
        import traceback
        st.code(traceback.format_exc())
